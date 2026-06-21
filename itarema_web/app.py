#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CONTROLE DE EXAMES LABORATORIAIS – PREFEITURA MUNICIPAL DE ITAREMA
Sistema Web – Contrato Nº 004/2023-SMS-02
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
import sqlite3, os, io, hashlib, secrets
from datetime import datetime, date, timedelta
from functools import wraps
from config import (MUNICIPIO_NOME, MUNICIPIO_UF, SECRETARIA_NOME, SISTEMA_SUBTITULO,
                    CONTRATO_NUM, CONTRATO_EMP, CONTRATO_VALOR, CONTRATO_ANO,
                    JUSTIFICATIVA_OPCOES, EXAMES_CONTRATO,
                    CONTRATO_INICIO, CONTRATO_VENCIMENTO)

app = Flask(__name__)
# IMPORTANTE: precisa ser FIXA (não gerada aleatoriamente a cada processo),
# senão cada worker do gunicorn assina a sessão com uma chave diferente e o
# usuário "perde" login/contrato selecionado ao cair em outro worker.
# Defina a variável de ambiente SECRET_KEY no Railway para um valor próprio em produção.
app.secret_key = os.environ.get("SECRET_KEY", "itarema-exames-lab-chave-fixa-c0ntrole-2026")

# ── Configurações de segurança da sessão ──────────────────────────────────────
app.config.update(
    SESSION_COOKIE_HTTPONLY  = True,
    SESSION_COOKIE_SAMESITE  = "Lax",
    SESSION_COOKIE_SECURE    = False,
    PERMANENT_SESSION_LIFETIME = timedelta(hours=8),
    # ── Flask-Mail (Gmail) ────────────────────────────────────────────────────
)

# Controle de tentativas de login (em memória)
_tentativas_login = {}   # {ip: {"count": N, "bloqueado_ate": datetime}}
MAX_TENTATIVAS   = 5     # máximo de erros antes de bloquear
BLOQUEIO_MINUTOS = 15    # tempo de bloqueio em minutos

DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "itarema_web.db"))

# ── Banco de dados ─────────────────────────────────────────────────────────────
def get_db():
    db_dir = os.path.dirname(os.path.abspath(DB_PATH))
    os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS exames (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL UNIQUE,
            descricao TEXT NOT NULL,
            tipo TEXT DEFAULT '',
            valor_unitario REAL DEFAULT 0.0,
            quantidade_contratada INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS autorizacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_autorizacao TEXT,
            data_autorizacao TEXT NOT NULL,
            nome_paciente TEXT NOT NULL,
            cpf_cns TEXT,
            codigo_exame TEXT NOT NULL,
            descricao_exame TEXT,
            quantidade_liberada INTEGER DEFAULT 1,
            valor_unitario REAL DEFAULT 0.0,
            valor_total REAL DEFAULT 0.0,
            responsavel TEXT,
            status TEXT DEFAULT 'AUTORIZADO',
            observacoes TEXT,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            criado_por_id INTEGER,
            criado_por_nome TEXT
        );
        CREATE TABLE IF NOT EXISTS orcamento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ano INTEGER NOT NULL UNIQUE,
            valor REAL DEFAULT 0.0
        );
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            login TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            perfil TEXT DEFAULT 'operador',
            ativo INTEGER DEFAULT 1,
            criado_em TEXT
        );
        CREATE TABLE IF NOT EXISTS reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT NOT NULL,
            token TEXT NOT NULL,
            expira_em TEXT NOT NULL,
            usado INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS contratos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'Laboratorial',
            empresa TEXT,
            valor_total REAL DEFAULT 0.0,
            data_inicio TEXT,
            data_vencimento TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS responsaveis_contrato (
            contrato_id INTEGER PRIMARY KEY,
            gestor_nome TEXT, gestor_cargo TEXT, gestor_cpf TEXT,
            gestor_telefone TEXT, gestor_email TEXT,
            fiscal_nome TEXT, fiscal_cargo TEXT, fiscal_cpf TEXT,
            fiscal_telefone TEXT, fiscal_email TEXT,
            atualizado_em TEXT
        );
    """)
    # Migração: adiciona colunas novas se ainda não existirem
    cols = [r[1] for r in conn.execute("PRAGMA table_info(autorizacoes)").fetchall()]
    if "criado_por_id"   not in cols: conn.execute("ALTER TABLE autorizacoes ADD COLUMN criado_por_id INTEGER")
    if "criado_por_nome" not in cols: conn.execute("ALTER TABLE autorizacoes ADD COLUMN criado_por_nome TEXT")
    if "justificativa"   not in cols: conn.execute("ALTER TABLE autorizacoes ADD COLUMN justificativa TEXT")
    ucols = [r[1] for r in conn.execute("PRAGMA table_info(usuarios)").fetchall()]
    if "email" not in ucols: conn.execute("ALTER TABLE usuarios ADD COLUMN email TEXT")

    # ── Multi-contrato: garante 1 contrato semente com os dados do config.py ────
    if conn.execute("SELECT COUNT(*) FROM contratos").fetchone()[0] == 0:
        conn.execute("""INSERT INTO contratos(numero,tipo,empresa,valor_total,data_inicio,data_vencimento,ativo)
                        VALUES(?,?,?,?,?,?,1)""",
                     (CONTRATO_NUM, "Laboratorial", CONTRATO_EMP, CONTRATO_VALOR, CONTRATO_INICIO, CONTRATO_VENCIMENTO))
    contrato_semente_id = conn.execute("SELECT id FROM contratos ORDER BY id LIMIT 1").fetchone()[0]

    # ── Migração: exames e orcamento passam a pertencer a um contrato ───────────
    ecols = [r[1] for r in conn.execute("PRAGMA table_info(exames)").fetchall()]
    if "contrato_id" not in ecols:
        conn.execute("ALTER TABLE exames RENAME TO exames_old")
        conn.execute("""CREATE TABLE exames (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contrato_id INTEGER NOT NULL,
            codigo TEXT NOT NULL,
            descricao TEXT NOT NULL,
            tipo TEXT DEFAULT '',
            valor_unitario REAL DEFAULT 0.0,
            quantidade_contratada INTEGER DEFAULT 0,
            UNIQUE(codigo, contrato_id)
        )""")
        conn.execute("""INSERT INTO exames(id,contrato_id,codigo,descricao,tipo,valor_unitario,quantidade_contratada)
                        SELECT id,?,codigo,descricao,tipo,valor_unitario,quantidade_contratada FROM exames_old""",
                     (contrato_semente_id,))
        conn.execute("DROP TABLE exames_old")

    ocols = [r[1] for r in conn.execute("PRAGMA table_info(orcamento)").fetchall()]
    if "contrato_id" not in ocols:
        conn.execute("ALTER TABLE orcamento RENAME TO orcamento_old")
        conn.execute("""CREATE TABLE orcamento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contrato_id INTEGER NOT NULL,
            ano INTEGER NOT NULL,
            valor REAL DEFAULT 0.0,
            UNIQUE(contrato_id, ano)
        )""")
        conn.execute("""INSERT INTO orcamento(id,contrato_id,ano,valor)
                        SELECT id,?,ano,valor FROM orcamento_old""", (contrato_semente_id,))
        conn.execute("DROP TABLE orcamento_old")

    acols = [r[1] for r in conn.execute("PRAGMA table_info(autorizacoes)").fetchall()]
    if "contrato_id" not in acols:
        conn.execute("ALTER TABLE autorizacoes ADD COLUMN contrato_id INTEGER")
        conn.execute("UPDATE autorizacoes SET contrato_id=? WHERE contrato_id IS NULL", (contrato_semente_id,))

    # ── Migração: responsaveis_contrato tinha PK fixa (id=1); passa a ser por contrato ──
    rcols = [r[1] for r in conn.execute("PRAGMA table_info(responsaveis_contrato)").fetchall()]
    if "contrato_id" not in rcols:
        conn.execute("ALTER TABLE responsaveis_contrato RENAME TO responsaveis_contrato_old")
        conn.execute("""CREATE TABLE responsaveis_contrato (
            contrato_id INTEGER PRIMARY KEY,
            gestor_nome TEXT, gestor_cargo TEXT, gestor_cpf TEXT,
            gestor_telefone TEXT, gestor_email TEXT,
            fiscal_nome TEXT, fiscal_cargo TEXT, fiscal_cpf TEXT,
            fiscal_telefone TEXT, fiscal_email TEXT,
            atualizado_em TEXT
        )""")
        old_rcols = [r[1] for r in conn.execute("PRAGMA table_info(responsaveis_contrato_old)").fetchall()]
        novos_campos_validos = ["gestor_nome","gestor_cargo","gestor_cpf","gestor_telefone","gestor_email",
                                 "fiscal_nome","fiscal_cargo","fiscal_cpf","fiscal_telefone","fiscal_email","atualizado_em"]
        # Só migra colunas que existem nos dois esquemas (versões antigas tinham
        # gestor_portaria/fiscal_portaria em vez de gestor_cpf/fiscal_cpf — não há de onde migrar esse valor).
        campos_comuns = [c for c in old_rcols if c in novos_campos_validos]
        if campos_comuns:
            campos_sql = ",".join(campos_comuns)
            conn.execute(f"""INSERT INTO responsaveis_contrato(contrato_id,{campos_sql})
                            SELECT ?,{campos_sql} FROM responsaveis_contrato_old""", (contrato_semente_id,))
        conn.execute("DROP TABLE responsaveis_contrato_old")

    if conn.execute("SELECT COUNT(*) FROM exames").fetchone()[0] == 0:
        conn.executemany(
            "INSERT OR IGNORE INTO exames(contrato_id,codigo,descricao,tipo,valor_unitario,quantidade_contratada) VALUES(?,?,?,?,?,?)",
            [(contrato_semente_id, *e) for e in EXAMES_CONTRATO])

    for row in conn.execute("SELECT id FROM contratos").fetchall():
        conn.execute("INSERT OR IGNORE INTO responsaveis_contrato(contrato_id) VALUES(?)", (row[0],))

    conn.execute("INSERT OR IGNORE INTO orcamento(contrato_id,ano,valor) VALUES(?,?,?)", (contrato_semente_id, CONTRATO_ANO, CONTRATO_VALOR))
    ano_atual = date.today().year
    conn.execute("INSERT OR IGNORE INTO orcamento(contrato_id,ano,valor) VALUES(?,?,?)", (contrato_semente_id, ano_atual, CONTRATO_VALOR))
    # Cria usuário admin padrão se não existir
    if conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] == 0:
        senha_hash = hashlib.sha256("admin123".encode()).hexdigest()
        conn.execute("INSERT INTO usuarios(nome,login,senha_hash,perfil,criado_em) VALUES(?,?,?,?,?)",
                     ("Administrador", "admin", senha_hash, "admin", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    # Permite reset de senha via variável de ambiente (use apenas para recuperação)
    reset_senha = os.environ.get("RESET_ADMIN_PASS", "")
    if reset_senha:
        nova_hash = hashlib.sha256(reset_senha.encode()).hexdigest()
        conn.execute("UPDATE usuarios SET senha_hash=? WHERE login='admin'", (nova_hash,))
    conn.commit(); conn.close()

# ── Utilitários ────────────────────────────────────────────────────────────────
def fmt_brl(v):
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",","X").replace(".",",").replace("X",".")
    except: return "R$ 0,00"

def parse_data(s):
    try: return datetime.strptime(s.strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
    except: return None

def fmt_data(s):
    try: return datetime.strptime(s.strip(), "%Y-%m-%d").strftime("%d/%m/%Y")
    except: return s or ""

def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def calc_prazo_contrato(contrato):
    """Calcula o percentual decorrido e status de vencimento do contrato informado
    (dict/Row com data_inicio e data_vencimento no formato YYYY-MM-DD)."""
    if not contrato:
        return None
    try:
        inicio = datetime.strptime(contrato["data_inicio"], "%Y-%m-%d").date()
        vencimento = datetime.strptime(contrato["data_vencimento"], "%Y-%m-%d").date()
    except Exception:
        return None

    hoje = date.today()
    total_dias = max((vencimento - inicio).days, 1)
    decorridos = (hoje - inicio).days
    pct = max(0, min(100, decorridos / total_dias * 100))

    dias_restantes = (vencimento - hoje).days
    meses_restantes = (vencimento.year - hoje.year) * 12 + (vencimento.month - hoje.month)
    if vencimento.day < hoje.day:
        meses_restantes -= 1

    if dias_restantes < 0:
        status, cor = "Vencido", "danger"
    elif meses_restantes < 6:
        status, cor = "Crítico", "danger"
    elif meses_restantes < 12:
        status, cor = "Atenção", "warning"
    else:
        status, cor = "Em dia", "success"

    return {
        "inicio": inicio, "vencimento": vencimento,
        "pct": pct, "dias_restantes": dias_restantes,
        "meses_restantes": meses_restantes,
        "status": status, "cor": cor,
    }

def get_contrato_ativo(db):
    """Retorna o contrato selecionado na sessão (ou o primeiro contrato ativo)."""
    cid = session.get("contrato_ativo_id")
    row = None
    if cid:
        row = db.execute("SELECT * FROM contratos WHERE id=? AND ativo=1", (cid,)).fetchone()
    if not row:
        row = db.execute("SELECT * FROM contratos WHERE ativo=1 ORDER BY id LIMIT 1").fetchone()
        if row:
            session["contrato_ativo_id"] = row["id"]
    return row

def get_ip():
    return request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()

def verificar_bloqueio(ip):
    """Retorna (bloqueado: bool, segundos_restantes: int)"""
    info = _tentativas_login.get(ip)
    if not info: return False, 0
    if info.get("bloqueado_ate") and datetime.now() < info["bloqueado_ate"]:
        restante = int((info["bloqueado_ate"] - datetime.now()).total_seconds())
        return True, restante
    return False, 0

def registrar_falha(ip):
    info = _tentativas_login.get(ip, {"count": 0, "bloqueado_ate": None})
    info["count"] += 1
    if info["count"] >= MAX_TENTATIVAS:
        info["bloqueado_ate"] = datetime.now() + timedelta(minutes=BLOQUEIO_MINUTOS)
        info["count"] = 0
    _tentativas_login[ip] = info

def limpar_tentativas(ip):
    _tentativas_login.pop(ip, None)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario_id"):
            flash("Faça login para acessar o sistema.", "warning")
            return redirect(url_for("login"))
        # Verifica timeout de sessão
        ultimo_acesso = session.get("ultimo_acesso")
        if ultimo_acesso:
            inativo = datetime.now() - datetime.fromisoformat(ultimo_acesso)
            if inativo > timedelta(hours=8):
                session.clear()
                flash("Sessão expirada por inatividade. Faça login novamente.", "warning")
                return redirect(url_for("login"))
        session["ultimo_acesso"] = datetime.now().isoformat()
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario_id"):
            return redirect(url_for("login"))
        if session.get("usuario_perfil") != "admin":
            flash("Acesso restrito a administradores.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

app.jinja_env.globals.update(
    fmt_brl=fmt_brl, fmt_data=fmt_data, hoje=date.today,
    MUNICIPIO_NOME=MUNICIPIO_NOME, MUNICIPIO_UF=MUNICIPIO_UF,
    SECRETARIA_NOME=SECRETARIA_NOME, SISTEMA_SUBTITULO=SISTEMA_SUBTITULO,
    JUST_OPCOES=JUSTIFICATIVA_OPCOES,
)

@app.context_processor
def inject_contrato_ativo():
    """Disponibiliza o contrato ativo (e a lista de contratos) em todos os templates."""
    db = get_db()
    if session.get("usuario_id"):
        contrato = get_contrato_ativo(db)
    else:
        contrato = db.execute("SELECT * FROM contratos WHERE ativo=1 ORDER BY id LIMIT 1").fetchone()
    lista = db.execute("SELECT id,numero,tipo,empresa FROM contratos WHERE ativo=1 ORDER BY tipo,numero").fetchall()
    db.close()
    return dict(
        CONTRATO_ATIVO=contrato,
        CONTRATOS_LISTA=lista,
        CONTRATO_NUM=contrato["numero"] if contrato else "",
        CONTRATO_EMP=contrato["empresa"] if contrato else "",
        CONTRATO_VALOR=contrato["valor_total"] if contrato else 0,
    )

# ── Cabeçalhos de segurança em todas as respostas ─────────────────────────────
@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"]        = "DENY"
    response.headers["X-XSS-Protection"]       = "1; mode=block"
    response.headers["Referrer-Policy"]        = "strict-origin-when-cross-origin"
    return response

# ── Rotas de Autenticação ──────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if session.get("usuario_id"):
        return redirect(url_for("dashboard"))
    ip = get_ip()
    bloqueado, segundos = verificar_bloqueio(ip)
    if bloqueado:
        minutos = segundos // 60 + 1
        flash(f"⛔ Muitas tentativas incorretas. Aguarde {minutos} minuto(s) para tentar novamente.", "danger")
        return render_template("login.html")
    if request.method == "POST":
        login_u = request.form.get("login","").strip()
        senha   = request.form.get("senha","").strip()
        if not login_u or not senha:
            flash("Preencha login e senha.", "danger")
            return render_template("login.html")
        db = get_db()
        user = db.execute("SELECT * FROM usuarios WHERE login=? AND ativo=1", (login_u,)).fetchone()
        db.close()
        if user and user["senha_hash"] == hash_senha(senha):
            limpar_tentativas(ip)
            session.permanent = True
            session["usuario_id"]      = user["id"]
            session["usuario_nome"]    = user["nome"]
            session["usuario_perfil"]  = user["perfil"]
            session["ultimo_acesso"]   = datetime.now().isoformat()
            flash(f"Bem-vindo, {user['nome']}!", "success")
            return redirect(url_for("dashboard"))
        registrar_falha(ip)
        info = _tentativas_login.get(ip, {})
        tentativas = MAX_TENTATIVAS - info.get("count", 0)
        if info.get("bloqueado_ate"):
            flash(f"⛔ Conta bloqueada por {BLOQUEIO_MINUTOS} minutos após múltiplas tentativas.", "danger")
        else:
            flash(f"Login ou senha incorretos. Tentativas restantes: {max(0,tentativas)}.", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Sessão encerrada.", "info")
    return redirect(url_for("login"))

@app.route("/recuperar-senha", methods=["GET","POST"])
def recuperar_senha():
    if request.method == "POST":
        login_u = request.form.get("login","").strip()
        db = get_db()
        user = db.execute("SELECT * FROM usuarios WHERE login=? AND ativo=1", (login_u,)).fetchone()
        # Sempre mostra a mesma mensagem para não revelar se o login existe
        if not user or not user["email"]:
            db.close()
            flash("Se o login existir e tiver e-mail cadastrado, você receberá o código em breve.", "info")
            return render_template("recuperar_senha.html")
        import random
        codigo = str(random.randint(100000, 999999))
        expira = (datetime.now() + timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")
        db.execute("UPDATE reset_tokens SET usado=1 WHERE login=?", (login_u,))
        db.execute("INSERT INTO reset_tokens(login,token,expira_em) VALUES(?,?,?)", (login_u, codigo, expira))
        db.commit(); db.close()
        # Envia o código por e-mail
        try:
            from sendgrid import SendGridAPIClient
            from sendgrid.helpers.mail import Mail as SGMail
            sg_msg = SGMail(
                from_email=os.environ.get("MAIL_FROM", "a.c.consultoriaemsaude2025@gmail.com"),
                to_emails=user["email"],
                subject="Recuperação de Senha – Sistema de Exames",
                plain_text_content=f"""Olá, {user["nome"]}!

Você solicitou a redefinição de senha no Sistema de Controle de Exames.

Seu código de verificação é: {codigo}

Este código é válido por 30 minutos.

Se você não solicitou essa redefinição, ignore este e-mail.

Atenciosamente,
Secretaria Municipal de Saúde de {MUNICIPIO_NOME}/{MUNICIPIO_UF}
"""
            )
            sg = SendGridAPIClient(os.environ.get("SENDGRID_API_KEY",""))
            sg.send(sg_msg)
        except BaseException as e:
            flash("Erro ao enviar e-mail. Contate o administrador.", "danger")
            return render_template("recuperar_senha.html")
        flash("Código enviado para o e-mail cadastrado. Verifique sua caixa de entrada.", "success")
        return render_template("recuperar_senha.html", login=login_u, aguardando_codigo=True)
    return render_template("recuperar_senha.html")

@app.route("/nova-senha", methods=["GET","POST"])
def nova_senha():
    if request.method == "POST":
        login_u = request.form.get("login","").strip()
        codigo  = request.form.get("codigo","").strip()
        nova    = request.form.get("nova_senha","").strip()
        conf    = request.form.get("confirmar_senha","").strip()
        if not nova or len(nova) < 6:
            flash("A senha deve ter pelo menos 6 caracteres.", "danger")
            return render_template("nova_senha.html")
        if nova != conf:
            flash("As senhas não coincidem.", "danger")
            return render_template("nova_senha.html")
        db = get_db()
        token = db.execute(
            "SELECT * FROM reset_tokens WHERE login=? AND token=? AND usado=0 AND expira_em > datetime('now','localtime')",
            (login_u, codigo)).fetchone()
        if not token:
            db.close()
            flash("Código inválido ou expirado. Solicite um novo código.", "danger")
            return render_template("nova_senha.html")
        db.execute("UPDATE usuarios SET senha_hash=? WHERE login=?", (hash_senha(nova), login_u))
        db.execute("UPDATE reset_tokens SET usado=1 WHERE id=?", (token["id"],))
        db.commit(); db.close()
        flash("✅ Senha redefinida com sucesso! Faça login com a nova senha.", "success")
        return redirect(url_for("login"))
    return render_template("nova_senha.html")

@app.route("/usuarios", methods=["GET","POST"])
@admin_required
def usuarios():
    db = get_db()
    if request.method == "POST":
        acao  = request.form.get("acao")
        nome  = request.form.get("nome","").strip()
        login_u = request.form.get("login","").strip()
        perfil  = request.form.get("perfil","operador")
        if acao == "add":
            senha = request.form.get("senha","").strip()
            email = request.form.get("email","").strip() or None
            if not nome or not login_u or not senha:
                flash("Preencha todos os campos.", "danger")
            else:
                try:
                    db.execute("INSERT INTO usuarios(nome,login,senha_hash,perfil,email,criado_em) VALUES(?,?,?,?,?,?)",
                               (nome, login_u, hash_senha(senha), perfil, email, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    db.commit()
                    flash(f"✅ Usuário '{login_u}' criado!", "success")
                except: flash(f"Login '{login_u}' já existe.", "danger")
        elif acao == "toggle":
            uid = request.form.get("uid")
            db.execute("UPDATE usuarios SET ativo = CASE WHEN ativo=1 THEN 0 ELSE 1 END WHERE id=?", (uid,))
            db.commit()
            flash("Status do usuário atualizado.", "info")
        elif acao == "update_email":
            uid   = request.form.get("uid")
            email = request.form.get("email","").strip() or None
            db.execute("UPDATE usuarios SET email=? WHERE id=?", (email, uid))
            db.commit()
            flash("✅ E-mail atualizado!", "success")
        elif acao == "reset_senha":
            uid   = request.form.get("uid")
            nova  = request.form.get("nova_senha","").strip()
            if nova:
                db.execute("UPDATE usuarios SET senha_hash=? WHERE id=?", (hash_senha(nova), uid))
                db.commit()
                flash("✅ Senha redefinida!", "success")
        elif acao == "del":
            uid = request.form.get("uid")
            if str(uid) == str(session.get("usuario_id")):
                flash("Não é possível excluir seu próprio usuário.", "danger")
            else:
                db.execute("DELETE FROM usuarios WHERE id=?", (uid,))
                db.commit()
                flash("Usuário excluído.", "warning")
        return redirect(url_for("usuarios"))
    rows = db.execute("SELECT * FROM usuarios ORDER BY nome").fetchall()
    db.close()
    return render_template("usuarios.html", rows=rows)

# ── Rotas Principais ───────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    ano = int(request.args.get("ano", date.today().year))
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    orc = (db.execute("SELECT valor FROM orcamento WHERE ano=? AND contrato_id=?", (ano, cid)).fetchone() or [CONTRATO_VALOR])[0]
    gasto = db.execute(
        "SELECT COALESCE(SUM(valor_total),0) FROM autorizacoes WHERE status!='CANCELADO' AND strftime('%Y',data_autorizacao)=? AND contrato_id=?",
        (str(ano), cid)).fetchone()[0]
    saldo = orc - gasto
    pct   = (gasto/orc*100) if orc > 0 else 0
    total_aut = db.execute("SELECT COUNT(*) FROM autorizacoes WHERE strftime('%Y',data_autorizacao)=? AND contrato_id=?", (str(ano), cid)).fetchone()[0]

    status_data = {}
    for st in ["AUTORIZADO","PENDENTE","REALIZADO","CANCELADO"]:
        r = db.execute("SELECT COUNT(*), COALESCE(SUM(valor_total),0) FROM autorizacoes WHERE status=? AND strftime('%Y',data_autorizacao)=? AND contrato_id=?", (st, str(ano), cid)).fetchone()
        status_data[st] = {"count": r[0], "valor": r[1]}

    ultimas = db.execute(
        "SELECT * FROM autorizacoes WHERE contrato_id=? ORDER BY id DESC LIMIT 10", (cid,)).fetchall()

    # Dados para gráfico mensal
    meses_labels, meses_valores = [], []
    for m in range(1, 13):
        v = db.execute(
            "SELECT COALESCE(SUM(valor_total),0) FROM autorizacoes WHERE strftime('%Y',data_autorizacao)=? AND strftime('%m',data_autorizacao)=? AND status!='CANCELADO' AND contrato_id=?",
            (str(ano), f"{m:02d}", cid)).fetchone()[0]
        meses_labels.append(["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][m-1])
        meses_valores.append(round(v, 2))
    db.close()
    prazo = calc_prazo_contrato(contrato)
    return render_template("dashboard.html", orc=orc, gasto=gasto, saldo=saldo, pct=pct,
                           total_aut=total_aut, status_data=status_data, ultimas=ultimas,
                           ano=ano, meses_labels=meses_labels, meses_valores=meses_valores,
                           prazo=prazo)

@app.route("/confirmar", methods=["GET","POST"])
@login_required
def confirmar_exames():
    # Prestador só acessa esta rota; admin/operador também podem ver
    if request.method == "POST":
        id_aut     = request.form.get("id")
        novo_status = request.form.get("status","").strip()
        if novo_status in ("REALIZADO","PENDENTE","CANCELADO"):
            db = get_db()
            db.execute("UPDATE autorizacoes SET status=? WHERE id=?", (novo_status, id_aut))
            db.commit(); db.close()
            flash(f"✅ Status atualizado para {novo_status}!", "success")
        return redirect(url_for("confirmar_exames",
                                paciente=request.args.get("paciente",""),
                                mes=request.args.get("mes",""),
                                exame=request.args.get("exame","")))
    db  = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    pac   = request.args.get("paciente","")
    mes   = request.args.get("mes","")
    exame = request.args.get("exame","")
    q  = "SELECT * FROM autorizacoes WHERE contrato_id=?"
    p  = [cid]
    if pac:   q += " AND nome_paciente LIKE ?";   p.append(f"%{pac}%")
    if mes:   q += " AND strftime('%m',data_autorizacao)=?"; p.append(mes.zfill(2))
    if exame: q += " AND (descricao_exame LIKE ? OR codigo_exame LIKE ?)"; p += [f"%{exame}%"]*2
    q += " ORDER BY data_autorizacao DESC, id DESC"
    rows = db.execute(q, p).fetchall()
    db.close()
    return render_template("confirmar.html", rows=rows,
                           filtros={"paciente":pac,"mes":mes,"exame":exame})

@app.route("/autorizacao", methods=["GET","POST"])
@login_required
def autorizacao():
    if session.get("usuario_perfil") == "prestador":
        flash("⛔ Acesso não permitido para o perfil Prestador.", "danger")
        return redirect(url_for("confirmar_exames"))
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    exames = db.execute("SELECT codigo, descricao, valor_unitario, quantidade_contratada FROM exames WHERE contrato_id=? ORDER BY descricao", (cid,)).fetchall()
    if request.method == "POST":
        data_db  = parse_data(request.form.get("data_autorizacao",""))
        pac      = request.form.get("nome_paciente","").strip()
        resp     = request.form.get("responsavel","").strip()
        cpf_cns       = request.form.get("cpf_cns","").strip() or None
        obs           = request.form.get("observacoes","").strip() or None
        justificativa = ",".join(request.form.getlist("justificativa")) or None
        # Gera número de autorização automático: ANO-SEQUENCIAL (ex: 2026-0001)
        ano_atual = date.today().year
        ultimo = db.execute(
            "SELECT numero_autorizacao FROM autorizacoes WHERE numero_autorizacao LIKE ? ORDER BY id DESC LIMIT 1",
            (f"{ano_atual}-%",)).fetchone()
        if ultimo and ultimo["numero_autorizacao"]:
            try: seq = int(ultimo["numero_autorizacao"].split("-")[1]) + 1
            except: seq = 1
        else:
            seq = 1
        num_aut = f"{ano_atual}-{seq:04d}"
        status_form = request.form.get("status","AUTORIZADO")

        # Validações dos campos comuns
        if not data_db:
            flash("Data inválida. Use o seletor de data.","danger")
            db.close(); return redirect(url_for("autorizacao"))
        if not pac:
            flash("Informe o nome do paciente.","danger")
            db.close(); return redirect(url_for("autorizacao"))
        if not resp:
            flash("Informe o responsável.","danger")
            db.close(); return redirect(url_for("autorizacao"))

        # Listas de exames submetidos pelo formulário
        codigos   = [c.strip().upper() for c in request.form.getlist("codigo_exame") if c.strip()]
        quantidades = request.form.getlist("quantidade_liberada")

        if not codigos:
            flash("Selecione pelo menos um exame.","danger")
            db.close(); return redirect(url_for("autorizacao"))

        # ── Valida cada exame antes de gravar qualquer um ──────────────────────
        erros = []
        itens = []  # lista de dicts prontos para inserção
        for i, cod in enumerate(codigos):
            try:
                qtde = int(quantidades[i]) if i < len(quantidades) else 1
                if qtde < 1: qtde = 1
            except:
                qtde = 1

            row_e = db.execute(
                "SELECT descricao, quantidade_contratada, valor_unitario FROM exames WHERE codigo=? AND contrato_id=?",
                (cod, cid)).fetchone()

            if not row_e:
                erros.append(f"Exame '{cod}' não encontrado no catálogo.")
                continue

            desc = row_e["descricao"]
            vu   = float(row_e["valor_unitario"])
            total_item = qtde * vu

            # Verificação de saldo (apenas para status diferente de CANCELADO)
            if status_form != "CANCELADO":
                qtd_contratada = int(row_e["quantidade_contratada"])
                qtd_usada = db.execute(
                    "SELECT COALESCE(SUM(quantidade_liberada),0) FROM autorizacoes WHERE codigo_exame=? AND status!='CANCELADO' AND contrato_id=?",
                    (cod, cid)).fetchone()[0]
                saldo_exame = qtd_contratada - qtd_usada
                if saldo_exame <= 0:
                    erros.append(f"❌ Saldo ESGOTADO para '{desc}' — sem unidades disponíveis no contrato.")
                    continue
                if qtde > saldo_exame:
                    erros.append(f"❌ Qtde solicitada ({qtde}) excede saldo disponível ({saldo_exame}) para '{desc}'.")
                    continue

            itens.append({
                "cod": cod, "desc": desc, "qtde": qtde,
                "vu": vu, "total": total_item
            })

        if erros:
            db.close()
            for e in erros:
                flash(e, "danger")
            return redirect(url_for("autorizacao"))

        # ── Grava um registro por exame ────────────────────────────────────────
        total_geral = 0.0
        for item in itens:
            db.execute("""INSERT INTO autorizacoes
                (numero_autorizacao,data_autorizacao,nome_paciente,cpf_cns,codigo_exame,
                 descricao_exame,quantidade_liberada,valor_unitario,valor_total,
                 responsavel,status,observacoes,justificativa,criado_por_id,criado_por_nome,contrato_id)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (num_aut, data_db, pac, cpf_cns,
                 item["cod"], item["desc"], item["qtde"], item["vu"], item["total"],
                 resp, status_form, obs, justificativa,
                 session.get("usuario_id"), session.get("usuario_nome"), cid))
            total_geral += item["total"]

        db.commit()
        qtd_exames = len(itens)
        plural = "exame" if qtd_exames == 1 else "exames"
        flash(f"✅ {qtd_exames} {plural} autorizado(s) para {pac} | Total: {fmt_brl(total_geral)}", "success")
        db.close()
        return redirect(url_for("autorizacao"))

    # Saldo financeiro
    ano_atual = date.today().year
    orc  = (db.execute("SELECT valor FROM orcamento WHERE ano=? AND contrato_id=?", (ano_atual, cid)).fetchone() or [CONTRATO_VALOR])[0]
    gasto = db.execute("SELECT COALESCE(SUM(valor_total),0) FROM autorizacoes WHERE status!='CANCELADO' AND strftime('%Y',data_autorizacao)=? AND contrato_id=?", (str(ano_atual), cid)).fetchone()[0]
    db.close()
    return render_template("autorizacao.html", exames=exames, orc=orc, gasto=gasto,
                           saldo=orc-gasto, hoje_str=date.today().strftime("%d/%m/%Y"))

@app.route("/api/exame/<codigo>")
@login_required
def api_exame(codigo):
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    row = db.execute("SELECT descricao,valor_unitario,quantidade_contratada FROM exames WHERE codigo=? AND contrato_id=?", (codigo.upper(), cid)).fetchone()
    usado = db.execute("SELECT COALESCE(SUM(quantidade_liberada),0) FROM autorizacoes WHERE codigo_exame=? AND status!='CANCELADO' AND contrato_id=?", (codigo.upper(), cid)).fetchone()[0]
    db.close()
    if not row: return jsonify({"erro": "Não encontrado"})
    return jsonify({"descricao": row["descricao"], "valor_unitario": row["valor_unitario"],
                    "qtd_contratada": row["quantidade_contratada"], "qtd_usada": int(usado),
                    "saldo": row["quantidade_contratada"] - int(usado)})

@app.route("/historico")
@login_required
def historico():
    db  = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    pac     = request.args.get("paciente","")
    st      = request.args.get("status","")
    mes     = request.args.get("mes","")
    exame   = request.args.get("exame","")
    cpf_cns = request.args.get("cpf_cns","").strip()
    q  = "SELECT * FROM autorizacoes WHERE contrato_id=?"
    p  = [cid]
    if pac:     q += " AND nome_paciente LIKE ?";   p.append(f"%{pac}%")
    if st:      q += " AND status=?";               p.append(st)
    if mes:     q += " AND strftime('%m',data_autorizacao)=?"; p.append(mes.zfill(2))
    if exame:   q += " AND (descricao_exame LIKE ? OR codigo_exame LIKE ?)"; p += [f"%{exame}%"]*2
    if cpf_cns: q += " AND cpf_cns LIKE ?"; p.append(f"%{cpf_cns.replace('.','').replace('-','').replace(' ','')}%")
    q += " ORDER BY data_autorizacao DESC, id DESC"
    rows = db.execute(q, p).fetchall()
    total_v = sum(r["valor_total"] for r in rows if r["status"] != "CANCELADO")
    db.close()
    return render_template("historico.html", rows=rows, total_v=total_v,
                           filtros={"paciente":pac,"status":st,"mes":mes,"exame":exame,"cpf_cns":cpf_cns},
                           uid_atual=int(session.get("usuario_id") or 0),
                           perfil_atual=session.get("usuario_perfil",""))

@app.route("/historico/excluir/<int:id>", methods=["POST"])
@login_required
def excluir_aut(id):
    db  = get_db()
    row = db.execute("SELECT criado_por_id FROM autorizacoes WHERE id=?", (id,)).fetchone()
    if not row:
        db.close()
        flash("Autorização não encontrada.", "danger")
        return redirect(url_for("historico"))
    eh_admin     = session.get("usuario_perfil") == "admin"
    uid_sessao   = int(session.get("usuario_id") or 0)
    uid_criador  = int(row["criado_por_id"]) if row["criado_por_id"] is not None else None
    eh_dono      = (uid_criador is None) or (uid_sessao == uid_criador)
    if not eh_admin and not eh_dono:
        db.close()
        flash("❌ Sem permissão! Só quem criou esta autorização ou um administrador pode excluí-la.", "danger")
        return redirect(url_for("historico"))
    db.execute("DELETE FROM autorizacoes WHERE id=?", (id,))
    db.commit(); db.close()
    flash("✅ Autorização excluída.", "warning")
    return redirect(url_for("historico"))

@app.route("/historico/editar/<int:id>", methods=["GET","POST"])
@login_required
def editar_aut(id):
    db = get_db()
    if request.method == "POST":
        data_db = parse_data(request.form.get("data_autorizacao",""))
        if not data_db: flash("Data inválida.","danger"); return redirect(url_for("editar_aut", id=id))
        just = ",".join(request.form.getlist("justificativa")) or None
        db.execute("""UPDATE autorizacoes SET numero_autorizacao=?,data_autorizacao=?,
                      nome_paciente=?,cpf_cns=?,responsavel=?,observacoes=?,status=?,justificativa=? WHERE id=?""",
            (request.form.get("numero_autorizacao","").strip() or None, data_db,
             request.form.get("nome_paciente","").strip(),
             request.form.get("cpf_cns","").strip() or None,
             request.form.get("responsavel","").strip() or None,
             request.form.get("observacoes","").strip() or None,
             request.form.get("status","AUTORIZADO"), just, id))
        db.commit(); db.close()
        flash("✅ Autorização atualizada!","success"); return redirect(url_for("historico"))
    row = db.execute("SELECT * FROM autorizacoes WHERE id=?", (id,)).fetchone()
    db.close()
    return render_template("editar.html", row=row)

@app.route("/saldo")
@login_required
def saldo_exames():
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    rows = db.execute("""
        SELECT e.codigo, e.descricao, e.tipo, e.valor_unitario, e.quantidade_contratada,
               COALESCE(SUM(CASE WHEN a.status!='CANCELADO' THEN a.quantidade_liberada ELSE 0 END),0) AS usada
        FROM exames e LEFT JOIN autorizacoes a ON a.codigo_exame=e.codigo AND a.contrato_id=e.contrato_id
        WHERE e.contrato_id=?
        GROUP BY e.codigo ORDER BY e.descricao
    """, (cid,)).fetchall()
    db.close()
    return render_template("saldo.html", rows=rows)

@app.route("/exames", methods=["GET","POST"])
@login_required
def exames():
    if session.get("usuario_perfil") == "prestador":
        flash("⛔ Acesso não permitido para o perfil Prestador.", "danger")
        return redirect(url_for("confirmar_exames"))
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    if request.method == "POST":
        acao = request.form.get("acao")
        cod  = request.form.get("codigo","").strip().upper()
        desc = request.form.get("descricao","").strip()
        tipo = request.form.get("tipo","").strip()
        try: val = float(request.form.get("valor_unitario","0").replace(",","."))
        except: val = 0.0
        try: qtd = int(request.form.get("quantidade_contratada","0"))
        except: qtd = 0
        if acao == "add":
            try:
                db.execute("INSERT INTO exames(contrato_id,codigo,descricao,tipo,valor_unitario,quantidade_contratada) VALUES(?,?,?,?,?,?)",(cid,cod,desc,tipo,val,qtd))
                flash(f"✅ Exame '{cod}' adicionado.","success")
            except: flash(f"Código '{cod}' já existe neste contrato.","danger")
        elif acao == "edit":
            cod_orig = request.form.get("codigo_original","")
            db.execute("UPDATE exames SET codigo=?,descricao=?,tipo=?,valor_unitario=?,quantidade_contratada=? WHERE codigo=? AND contrato_id=?",
                       (cod,desc,tipo,val,qtd,cod_orig,cid))
            flash(f"✅ Exame '{cod}' atualizado.","success")
        elif acao == "del":
            if session.get("usuario_perfil") != "admin":
                flash("⛔ Apenas administradores podem excluir exames.", "danger")
            else:
                db.execute("DELETE FROM exames WHERE codigo=? AND contrato_id=?", (cod, cid))
                flash(f"Exame '{cod}' excluído.","warning")
        db.commit()
        return redirect(url_for("exames"))
    rows = db.execute("SELECT * FROM exames WHERE contrato_id=? ORDER BY codigo", (cid,)).fetchall()
    db.close()
    return render_template("exames.html", rows=rows)

@app.route("/configuracoes/responsaveis", methods=["GET","POST"])
@login_required
def gestor_fiscal():
    if session.get("usuario_perfil") == "prestador":
        flash("⛔ Acesso não permitido para o perfil Prestador.", "danger")
        return redirect(url_for("confirmar_exames"))
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else None
    campos = ["gestor_nome","gestor_cargo","gestor_cpf","gestor_telefone","gestor_email",
              "fiscal_nome","fiscal_cargo","fiscal_cpf","fiscal_telefone","fiscal_email"]
    if request.method == "POST":
        valores = [request.form.get(c,"").strip() or None for c in campos]
        sets = ", ".join(f"{c}=?" for c in campos)
        db.execute("INSERT OR IGNORE INTO responsaveis_contrato(contrato_id) VALUES(?)", (cid,))
        db.execute(f"UPDATE responsaveis_contrato SET {sets}, atualizado_em=? WHERE contrato_id=?",
                   (*valores, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cid))
        db.commit(); db.close()
        flash("✅ Dados de Gestor e Fiscal do contrato atualizados!", "success")
        return redirect(url_for("gestor_fiscal"))
    row = db.execute("SELECT * FROM responsaveis_contrato WHERE contrato_id=?", (cid,)).fetchone()
    db.close()
    return render_template("gestor_fiscal.html", row=row)

@app.route("/contrato/selecionar/<int:cid>", methods=["POST"])
@login_required
def selecionar_contrato(cid):
    db = get_db()
    achado = db.execute("SELECT id, numero, ativo FROM contratos WHERE id=?", (cid,)).fetchone()
    db.close()
    if achado and achado["ativo"] == 1:
        session["contrato_ativo_id"] = cid
        flash(f"✅ Contrato '{achado['numero']}' selecionado.", "success")
    elif achado:
        flash(f"Contrato inválido (id={cid}, numero={achado['numero']}, ativo={achado['ativo']!r}).", "danger")
    else:
        flash(f"Contrato inválido (id={cid} não encontrado na tabela contratos).", "danger")
    voltar = request.referrer or url_for("dashboard")
    return redirect(voltar)

@app.route("/configuracoes/contratos", methods=["GET","POST"])
@admin_required
def contratos():
    db = get_db()
    if request.method == "POST":
        acao = request.form.get("acao")
        if acao == "add":
            numero  = request.form.get("numero","").strip()
            tipo    = request.form.get("tipo","").strip() or "Outro"
            empresa = request.form.get("empresa","").strip() or None
            try: valor = float(request.form.get("valor_total","0").replace(",","."))
            except: valor = 0.0
            inicio     = request.form.get("data_inicio","").strip() or None
            vencimento = request.form.get("data_vencimento","").strip() or None
            if not numero or not inicio or not vencimento:
                flash("Informe número, data de início e vencimento do contrato.", "danger")
            else:
                cur = db.execute("""INSERT INTO contratos(numero,tipo,empresa,valor_total,data_inicio,data_vencimento,ativo)
                                    VALUES(?,?,?,?,?,?,1)""", (numero, tipo, empresa, valor, inicio, vencimento))
                novo_id = cur.lastrowid
                db.execute("INSERT OR IGNORE INTO orcamento(contrato_id,ano,valor) VALUES(?,?,?)", (novo_id, date.today().year, valor))
                db.execute("INSERT OR IGNORE INTO responsaveis_contrato(contrato_id) VALUES(?)", (novo_id,))
                flash(f"✅ Contrato '{numero}' ({tipo}) cadastrado!", "success")
        elif acao == "edit":
            cidf    = request.form.get("cid")
            numero  = request.form.get("numero","").strip()
            tipo    = request.form.get("tipo","").strip() or "Outro"
            empresa = request.form.get("empresa","").strip() or None
            try: valor = float(request.form.get("valor_total","0").replace(",","."))
            except: valor = 0.0
            inicio     = request.form.get("data_inicio","").strip() or None
            vencimento = request.form.get("data_vencimento","").strip() or None
            db.execute("""UPDATE contratos SET numero=?,tipo=?,empresa=?,valor_total=?,data_inicio=?,data_vencimento=?
                          WHERE id=?""", (numero, tipo, empresa, valor, inicio, vencimento, cidf))
            flash(f"✅ Contrato '{numero}' atualizado.", "success")
        elif acao == "toggle":
            cidf = request.form.get("cid")
            db.execute("UPDATE contratos SET ativo = CASE WHEN ativo=1 THEN 0 ELSE 1 END WHERE id=?", (cidf,))
            flash("Status do contrato alterado.", "info")
        db.commit(); db.close()
        return redirect(url_for("contratos"))
    rows = db.execute("SELECT * FROM contratos ORDER BY ativo DESC, tipo, numero").fetchall()
    db.close()
    return render_template("contratos.html", rows=rows)

@app.route("/relatorio")
@login_required
def relatorio():
    ano = int(request.args.get("ano", CONTRATO_ANO))
    db  = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    orc = (db.execute("SELECT valor FROM orcamento WHERE ano=? AND contrato_id=?", (ano, cid)).fetchone() or [CONTRATO_VALOR])[0]
    meses_nomes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                   "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    dados = []
    acum = 0.0
    for m, nome in enumerate(meses_nomes, 1):
        q, v = db.execute("""SELECT COUNT(*), COALESCE(SUM(valor_total),0) FROM autorizacoes
            WHERE strftime('%Y',data_autorizacao)=? AND strftime('%m',data_autorizacao)=? AND status!='CANCELADO' AND contrato_id=?""",
            (str(ano), f"{m:02d}", cid)).fetchone()
        acum += v
        dados.append({"mes": nome, "qtd": q, "valor": v, "acumulado": acum,
                      "pct": (v/orc*100) if orc > 0 else 0})
    # Contagem por justificativa
    JUST_OPCOES = ["DIABÉTICO","HIPERTENSO","GESTANTE","GESTANTE DE ALTO RISCO","INTERNADO","PÚBLICO GERAL"]
    just_contagem = {}
    todas = db.execute(
        "SELECT justificativa FROM autorizacoes WHERE strftime('%Y',data_autorizacao)=? AND status!='CANCELADO' AND contrato_id=?",
        (str(ano), cid)).fetchall()
    for opcao in JUST_OPCOES:
        just_contagem[opcao] = sum(1 for r in todas if r["justificativa"] and opcao in r["justificativa"].split(","))
    db.close()
    return render_template("relatorio.html", dados=dados, ano=ano, orc=orc,
                           anos=list(range(2022, 2031)),
                           just_contagem=just_contagem, just_opcoes=JUST_OPCOES)

@app.route("/limpar-base", methods=["POST"])
@admin_required
def limpar_base():
    db = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    db.execute("DELETE FROM autorizacoes WHERE contrato_id=?", (cid,))
    db.commit(); db.close()
    flash("🗑️ Base de autorizações deste contrato limpa com sucesso! Pronto para uso real.", "warning")
    return redirect(url_for("historico"))

@app.route("/orcamento", methods=["POST"])
@login_required
def set_orcamento():
    try:
        ano = int(request.form.get("ano", CONTRATO_ANO))
        val = float(request.form.get("valor","0").replace(",","."))
        db  = get_db()
        contrato = get_contrato_ativo(db)
        cid = contrato["id"] if contrato else 0
        db.execute("INSERT INTO orcamento(contrato_id,ano,valor) VALUES(?,?,?) ON CONFLICT(contrato_id,ano) DO UPDATE SET valor=excluded.valor", (cid,ano,val))
        db.commit(); db.close()
        flash(f"✅ Orçamento {ano} definido: {fmt_brl(val)}","success")
    except: flash("Erro ao salvar orçamento.","danger")
    return redirect(url_for("dashboard"))

@app.route("/exportar/<int:ano>")
@login_required
def exportar(ano):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("openpyxl não instalado.","danger"); return redirect(url_for("relatorio"))
    db   = get_db()
    contrato = get_contrato_ativo(db)
    cid = contrato["id"] if contrato else 0
    rows = db.execute("""SELECT data_autorizacao,numero_autorizacao,nome_paciente,cpf_cns,
                                codigo_exame,descricao_exame,quantidade_liberada,valor_unitario,
                                valor_total,responsavel,status,observacoes,justificativa
                         FROM autorizacoes WHERE strftime('%Y',data_autorizacao)=? AND contrato_id=?
                         ORDER BY data_autorizacao,id""", (str(ano), cid)).fetchall()
    db.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Autorizações {ano}"
    hdrs = ["Data","Nº Aut.","Paciente","CPF/CNS","Código","Descrição Exame","Qtde","Val.Unit","Val.Total","Responsável","Status","Obs","Justificativa"]
    for c, h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True,color="FFFFFF",name="Arial")
        cell.fill = PatternFill("solid",fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    for ri, r in enumerate(rows,2):
        vals = [fmt_data(r[0]),r[1] or "",r[2],r[3] or "",r[4],r[5] or "",r[6],r[7],r[8],r[9] or "",r[10] or "",r[11] or "",r[12] or ""]
        for ci, v in enumerate(vals,1):
            cell = ws.cell(ri,ci,v)
            if ci in (8,9): cell.number_format = 'R$ #,##0.00'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"Itarema_Exames_{ano}.xlsx",
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Inicializa o banco sempre que o módulo for carregado (gunicorn ou direto)
init_db()

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
